from itertools import product
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
import numpy as np
import pandas as pd

# ABC
def ABCAnalysisDisplay(wb, data):
    ws = wb.create_sheet(title="ABC_VEN_Result")
    Years = data["NĂM"].unique()
    
    InitialRow = 1
    InitialColumn = 1
    
    AllDataColumn = 2
    AllDataRow = 18
    AllDataRowStep = 18
    AllDataRowHeader = 17
    
    for y in Years:
        dataByYear = data[data["NĂM"] == y]
        ws.cell(row = InitialRow, column = InitialColumn, value = y)
        
        ws.cell(row = InitialRow, column = InitialColumn + 1, value = "Giá trị")
        ws.cell(row = InitialRow, column = InitialColumn + 2, value = "% Giá trị")
        ws.cell(row = InitialRow, column = InitialColumn + 3, value = "Số lượng")
        ws.cell(row = InitialRow, column = InitialColumn + 4, value = "% Số lượng")
        
        # ABC
        AddCellValue_ABCVEN(ws, row=InitialRow + 1, column=InitialColumn, 
                     columnHeader='A', dataCol='ABC', dataColVal='A', data=dataByYear,
                     AllDataRow=AllDataRow, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)
        AddCellValue_ABCVEN(ws, row=InitialRow + 2, column=InitialColumn, 
                     columnHeader='B', dataCol='ABC', dataColVal='B', data=dataByYear,
                    AllDataRow=AllDataRow+1, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)
        
        AddCellValue_ABCVEN(ws, row=InitialRow + 3, column=InitialColumn, 
                     columnHeader='C', dataCol='ABC', dataColVal='C', data=dataByYear,
                    AllDataRow=AllDataRow+2, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)

        # VEN
        AddCellValue_ABCVEN(ws, row=InitialRow + 4, column=InitialColumn, 
                     columnHeader='V', dataCol='VEN', dataColVal='V', data=dataByYear,
                    AllDataRow=AllDataRow+3, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)

        AddCellValue_ABCVEN(ws, row=InitialRow + 5, column=InitialColumn, 
                     columnHeader='E', dataCol='VEN', dataColVal='E', data=dataByYear,
                    AllDataRow=AllDataRow+4, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)

        AddCellValue_ABCVEN(ws, row=InitialRow + 6, column=InitialColumn, 
                     columnHeader='N', dataCol='VEN', dataColVal='N', data=dataByYear, 
                    AllDataRow=AllDataRow+5, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                    AllDataRowHeader=AllDataRowHeader, year=y)

        # ABC-VEN
        for ind, val in enumerate(product(['A', 'B', 'C'], ['V', 'E', 'N'])):
            AddCellValue_ABCVEN(ws, row=InitialRow + 7 + ind, column=InitialColumn, 
                         columnHeader=''.join(val), dataCol='ABC-VEN', dataColVal=''.join(val), data=dataByYear,
                        AllDataRow=AllDataRow+6+ind, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                        AllDataRowHeader=AllDataRowHeader, year=y)
        
        InitialColumn += 8
        AllDataColumn += 1
        
    ChartDraw(ws, ChartName='Giá trị thuốc theo ABC/VEN qua các năm', 
              MinRow=17, MinColumn=2, MaxRow=32, MaxColumn=2+len(Years)-1, 
              CatsCol=1, CatsMinRow=18, CatsMaxRow=32, 
              ChartPos='J18', xTitle='Năm', yTitle='Giá trị')
    
    ChartDraw(ws, ChartName='Tỷ lệ giá trị thuốc theo ABC/VEN qua các năm', 
              MinRow=17+AllDataRowStep, MinColumn=2, MaxRow=32+AllDataRowStep, MaxColumn=2+len(Years)-1, 
              CatsCol=1, CatsMinRow=18+AllDataRowStep, CatsMaxRow=32+AllDataRowStep, 
              ChartPos=''.join(['J', str(18+AllDataRowStep)]), xTitle='Năm', yTitle='Tỷ lệ')
    
    
    ChartDraw(ws, ChartName='Số lượng thuốc theo ABC/VEN qua các năm', 
              MinRow=17+AllDataRowStep*2, MinColumn=2, MaxRow=32+AllDataRowStep*2, MaxColumn=2+len(Years)-1, 
              CatsCol=1, CatsMinRow=18+AllDataRowStep*2, CatsMaxRow=32+AllDataRowStep*2, 
              ChartPos=''.join(['J', str(18+AllDataRowStep*2)]), xTitle='Năm', yTitle='Số lượng')
   
    ChartDraw(ws, ChartName='Tỷ lệ số lượng thuốc theo ABC/VEN qua các năm', 
                  MinRow=17+AllDataRowStep*3, MinColumn=2, MaxRow=32+AllDataRowStep*3, MaxColumn=2+len(Years)-1, 
                  CatsCol=1, CatsMinRow=18+AllDataRowStep*3, CatsMaxRow=32+AllDataRowStep*3, 
                  ChartPos=''.join(['J', str(18+AllDataRowStep*3)]), xTitle='Năm', yTitle='Tỷ lệ')
    
    return
    
def GetDataResult_ABCVEN(data, subColumn, subValue): # value, number and %
    SubData = data[data[subColumn] == subValue]
    ValuePercentage = np.round(SubData['THÀNH TIỀN'].sum() / data['THÀNH TIỀN'].sum() * 100, decimals=2)
    NumberPercentage = np.round(SubData.shape[0] / data.shape[0] * 100, decimals=2)
    return SubData['THÀNH TIỀN'].sum(), ValuePercentage, SubData.shape[0], NumberPercentage

def AddDataResult_ABCVEN(ws, row, column, columnHeader, results,
                     AllDataRow, AllDataColumn, AllDataRowStep, AllDataRowHeader):
    for i in range(4):
        ws.cell(row = row, column = column + i, value = results[i])
        ws.cell(row = AllDataRow+(AllDataRowStep*(i)), column = AllDataColumn, value = results[i])
        ws.cell(row = AllDataRow+(AllDataRowStep*(i)), column=1, value=columnHeader)
        
def AddCellValue_ABCVEN(ws, row, column,  
                 columnHeader, dataCol, dataColVal, data,
                    AllDataRow, AllDataColumn, AllDataRowStep, AllDataRowHeader, year):
    ws.cell(row=row, column=column, value=columnHeader)
    
    if not ws.cell(row=AllDataRowHeader, column=AllDataColumn).value:
        ws.cell(row=AllDataRowHeader, column=AllDataColumn, value=year)
        ws.cell(row=AllDataRowHeader+AllDataRowStep, column=AllDataColumn, value=year)
        
    AddDataResult_ABCVEN(ws, row=row, column=column+1, columnHeader=columnHeader, 
                      results=GetDataResult_ABCVEN(data, dataCol, dataColVal),
                      AllDataRow=AllDataRow, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep, 
                      AllDataRowHeader=AllDataRowHeader)

# XYZ
def XYZAnalysisDisplay(wb, data):
    ws = wb.create_sheet(title="XYZ_Result")
    Years = data["NĂM"].unique()
    
    InitialRow = 1
    InitialColumn = 1
    
    AllDataColumn = 2
    AllDataRow = 8
    AllDataRowStep = 5
    AllDataRowHeader = 7
    
    for y in Years:
        dataByYear = data[data["NĂM"] == y]
        ws.cell(row = InitialRow, column = InitialColumn, value = y)
        
        ws.cell(row = InitialRow, column = InitialColumn + 1, value = "Số lượng")
        ws.cell(row = InitialRow, column = InitialColumn + 2, value = "% Số lượng")
        
        # XYZ
        AddCellValue_XYZ(ws, row=InitialRow + 1, column=InitialColumn, 
                     columnHeader='X', dataCol='XYZ', dataColVal='X', data=dataByYear, 
                         AllDataRow=AllDataRow, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                        AllDataRowHeader=AllDataRowHeader, year=y)

        AddCellValue_XYZ(ws, row=InitialRow + 2, column=InitialColumn, 
                     columnHeader='Y', dataCol='XYZ', dataColVal='Y', data=dataByYear, 
                         AllDataRow=AllDataRow+1, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                        AllDataRowHeader=AllDataRowHeader, year=y)
        
        AddCellValue_XYZ(ws, row=InitialRow + 3, column=InitialColumn, 
                     columnHeader='Z', dataCol='XYZ', dataColVal='Z', data=dataByYear, 
                         AllDataRow=AllDataRow+2, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep,
                        AllDataRowHeader=AllDataRowHeader, year=y)

        InitialColumn += 4
        AllDataColumn += 1
    
    ChartDraw(ws, ChartName='Số lượng thuốc theo XYZ qua các năm', 
              MinRow=7, MinColumn=2, MaxRow=10, MaxColumn=2+len(Years)-1, 
              CatsCol=1, CatsMinRow=8, CatsMaxRow=10, 
              ChartPos='J8', xTitle='Năm', yTitle='Số lượng')
    
    ChartDraw(ws, ChartName='Tỷ lệ thuốc theo XYZ qua các năm', 
              MinRow=7+AllDataRowStep, MinColumn=2, MaxRow=10+AllDataRowStep, MaxColumn=2+len(Years)-1, 
              CatsCol=1, CatsMinRow=8+AllDataRowStep, CatsMaxRow=10+AllDataRowStep, 
              ChartPos=''.join(['J', str(8+20)]), xTitle='Năm', yTitle='Tỷ lệ')
        
    return

def GetDataResult_XYZ(data, subColumn, subValue): # value, number and %
    SubData = data[data[subColumn] == subValue]
    NumberPercentage = np.round(SubData.shape[0] / data.shape[0] * 100, decimals=2)
    return SubData.shape[0], NumberPercentage

def AddDataResult_XYZ(ws, row, column, columnHeader, results, 
                     AllDataRow, AllDataColumn, AllDataRowStep, AllDataRowHeader):
    for i in range(2):
        ws.cell(row = row, column = column + i, value = results[i])
        ws.cell(row = AllDataRow+(AllDataRowStep*(i)), column = AllDataColumn, value = results[i])
        ws.cell(row = AllDataRow+(AllDataRowStep*(i)), column=1, value=columnHeader)
        
def AddCellValue_XYZ(ws, row, column, 
                 columnHeader, dataCol, dataColVal, data, 
                    AllDataRow, AllDataColumn, AllDataRowStep, AllDataRowHeader, year):
    ws.cell(row=row, column=column, value=columnHeader)
    
    if not ws.cell(row=AllDataRowHeader, column=AllDataColumn).value:
        ws.cell(row=AllDataRowHeader, column=AllDataColumn, value=year)
        ws.cell(row=AllDataRowHeader+AllDataRowStep, column=AllDataColumn, value=year)
    
    AddDataResult_XYZ(ws, row=row, column=column+1, columnHeader=columnHeader, 
                      results=GetDataResult_XYZ(data, dataCol, dataColVal), 
                      AllDataRow=AllDataRow, AllDataColumn=AllDataColumn, AllDataRowStep=AllDataRowStep, 
                      AllDataRowHeader=AllDataRowHeader)

def ChartDraw(ws, ChartName, 
              MinRow, MinColumn, MaxRow, MaxColumn,
             CatsCol, CatsMinRow, CatsMaxRow,
             ChartPos, xTitle, yTitle):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = ChartName
    chart.y_axis.title = yTitle
    chart.x_axis.title = xTitle
    
    data = Reference(ws, min_row=MinRow, min_col=MinColumn, max_row=MaxRow, max_col=MaxColumn)
    cats = Reference(ws, min_row=CatsMinRow, max_row=CatsMaxRow, min_col=CatsCol)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, ChartPos)